import openpyxl
import os
import sys
import time
import win32com.client
from openpyxl.utils import get_column_letter
import numpy as np
from fgdb_reader import Read_FGDB



class Edit_Excel(object):
    """
    Class contains methods for editing an Microsoft Excel file.
    
    Attributes
    ----------
    filename : str
        Path to the Excel-file
    
    """
    
    def __init__(self, filename):
        self.filename = filename


    def insert_cell_block(self, upper_left, cell_block, header=0):
        """
        Overwrite a cell-block in an Excel-file.
        
        Parameters
        ----------
        sh_name : str
            Name of the worksheet where the cells will 
        upper_left : list of int
            Indices of the upper-left cell that will be overwritten.
            Cell A1 equals upper_left = [1, 1].
        cell_block : Matix of str or float or int
            Cell-block that will be copied to the Excel-file.
        header : int, optional
            Number of ignored head rows.
        """
        cell_block = np.delete(cell_block, (header), axis=0)
        wb = openpyxl.load_workbook(self.filename)
        ws1 = wb.active
        ws1.title = "sheet1"
        block_rows, block_cols = np.shape(cell_block)
        start_row = upper_left[0]
        start_col = upper_left[1]
        for row in range(block_rows):
            for col in range(block_cols):
                cell_id = get_column_letter(start_col+col) + str(start_row+row)
                ws1[cell_id] = cell_block[row, col]
        wb.save(self.filename)
        wb.close()
    
    
    
    def append_rows(self, rows):
        """
        Write and save new lines in .xlsx file.
        
        Parameters
        ----------
        rows : np.array of str or int or float
            Rows that will be appended to the Excel-file
            
        """
        wb = openpyxl.load_workbook(self.filename)
        ws1 = wb.active
        ws1.title = "range names"
        for i in range(len(rows[:, 1])):
            ws1.append(rows[i, :])
        wb.save(self.filename)
        wb.close()
        
              
    def clear_cellblock(self, upper_left, lower_right):
        """
        Clear all Cells between upper_left and lower_right corner.
        
        Parameters
        ----------
        upper_left : str
            Marks the upper-left corner of the cellblock.
            Example: 'A1'
        lower_right : str
            Marks the lower-right corner of the cellblock.
            Example: 'D90'
            
        """
        excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = 0
        wb = excel.Workbooks.Open(self.filename)
        excel.Range(upper_left+':'+lower_right).Select()
        excel.Selection.ClearContents()
        wb.Save()
        wb.Close()
        
        
    def delete_rows(self, first, last):
        """
        Delete all rows from first to last.
        
        Parameters
        ----------
        first : int
            first row to delete
        last : int
            last row to delete
            
        """
        excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = 0
        wb = excel.Workbooks.Open(self.filename)
        sh = wb.ActiveSheet
        rangeObj = sh.Range('A'+str(first)+':A'+str(last))
        rangeObj.EntireRow.Delete()    
        wb.Save()
        wb.Close() 
        
        
    def delete_columns(self, first, last):
        """
        Delete all columns from first to last.
        
        Parameters
        ----------
        first : int
            first row to delete
        last : int
            last row to delete
         
        """
        excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = 0
        wb = excel.Workbooks.Open(self.filename)
        sh = wb.ActiveSheet
        rangeObj = sh.Range(first+'1:'+last+'1')
        rangeObj.EntireColumn.Delete()    
        wb.Save()
        wb.Close() 
        
        
    def open_file_in_excel(self):
        """
        Open .xlsx file in Excel and wait for 5 seconds to check if saving works
        """
        # open demo.xlsx file in Excel
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = True
        xl.Workbooks.Open(self.filename)
    
    
    
    def check_if_file_is_open(self):
        """
        Check if a file is opened in Excel.
        If file is open: Save and close
        """
        excel = win32com.client.Dispatch("Excel.Application")
        if excel.Workbooks.Count > 0:
            for i in range(1, excel.Workbooks.Count+1):
                #print(excel.Workbooks.Item(i).Name, filename)
                if excel.Workbooks.Item(i).Name == self.filename:
                    wb = excel.Workbooks.Item(i)
                    wb.Save()
                    wb.Close()
                    break    

        
        
if __name__ == '__main__':        
    
    
    # Variables:
    filename = 'demo.xlsx'
    filename2 = r'C:\ProjektCheck\test_excel\demo.xlsx'
    gdb_path = r"C:\ProjektCheck\3 Benutzerdefinierte Projekte\test_mean" + \
        "\FGDB_Definition_Projekt.gdb"
    test_data = 'Wohnen_WE_in_Gebaeudetypen'
    columns = np.array(["WE", "EW_je_WE"])
    
    Results = Read_FGDB(gdb_path, test_data, columns, 'WE>0 and WE<20')
    Results.get_result_block()
    Results.sort_results_by_col(1)
    cell_block = Results.result_block
    
    Excel_Editor = Edit_Excel(filename2)
    Excel_Editor.insert_cell_block([4, 8], cell_block)
    Excel_Editor.open_file_in_excel()
    
    print('done')
    
    
    
    
    
    """
    MAY BE USEFUL LATER:
    
    try:
        ws1.append(("after opening excel", "after opening excel"))
        wb.save(filename)
    except:
        print('file is already opened')
    # not used: opens demo.xlsx file
    #os.system('start excel.exe "%s\\demo.xlsx"' % (sys.path[0], ))
    """
    
    








